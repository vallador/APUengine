from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.core.apu_calculator import APUCalculator
from app.modules.engineering_module import EngineeringModule
from app.modules.pricing_module import PricingModule
from app.database.db_manager import engine_db

from app.core.config import EXCEL_PALETTE

# ── Colour palette (Corporate Green) ──────────────────────────
C_HEADER  = EXCEL_PALETTE['header']
C_MAT     = EXCEL_PALETTE['material']
C_EQ      = EXCEL_PALETTE['equipment']
C_MO      = EXCEL_PALETTE['labor']
C_TOTAL   = EXCEL_PALETTE['total']
C_AIU     = EXCEL_PALETTE['aiu_row']

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_cell(ws, row, col, value, bold=False, size=10, bg=None, align="left", border=None, number_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=size, color="FFFFFF" if bg == C_HEADER else "000000")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = border
    if number_fmt:
        c.number_format = number_fmt
    return c


class ExcelGenerator:
    @staticmethod
    def generate_apu_report(output_path="APU_Report.xlsx", variant_ids=None, estimation_items=None, project_name="INGENIERÍA INTEGRAL", project_nit=""):
        wb = Workbook()
        ws_index = wb.active
        ws_index.title = "Índice"

        # ── Resolve variants ──────────────────────────────────────
        if variant_ids:
            placeholders = ','.join(['?'] * len(variant_ids))
            variants = engine_db.execute_query(
                f"SELECT vp.id, vp.nombre, p.nombre, p.unidad_base, p.categoria, vp.especificacion_tecnica "
                f"FROM variantes_proceso vp JOIN procesos p ON vp.proceso_id = p.id "
                f"WHERE vp.id IN ({placeholders})", variant_ids)
        else:
            variants = engine_db.execute_query(
                "SELECT vp.id, vp.nombre, p.nombre, p.unidad_base, p.categoria, vp.especificacion_tecnica "
                "FROM variantes_proceso vp JOIN procesos p ON vp.proceso_id = p.id")

        # ── AIU factors ──────────────────────────────────────────
        aiu = APUCalculator.get_aiu_factors()

        # ── Index sheet header ────────────────────────────────────
        ws_index.column_dimensions['A'].width = 8
        ws_index.column_dimensions['B'].width = 40
        ws_index.column_dimensions['C'].width = 12
        ws_index.column_dimensions['D'].width = 14
        ws_index.column_dimensions['E'].width = 16
        ws_index.row_dimensions[1].height = 30
        _set_cell(ws_index, 1, 1, "LISTADO DE ANÁLISIS DE PRECIOS UNITARIOS",
                  bold=True, size=13, bg=C_HEADER, align="center")
        ws_index.merge_cells("A1:E1")
        for col, hdr in enumerate(["#", "Actividad", "Unidad", "V. Técnico", "V. Comercial"], start=1):
            _set_cell(ws_index, 2, col, hdr, bold=True, bg=C_HEADER, align="center")

        idx_row = 3
        grand_total = 0.0
        item_counter = 1

        for var in (variants or []):
            v_id, v_name, p_name, unit, categoria, v_desc = var
            qty = 1.0
            # Try to get qty from estimation_items if supplied
            if estimation_items:
                match = next((e for e in estimation_items if e['id'] == v_id), None)
                if match: qty = match.get('qty', 1.0)

            # ── Build APU detail sheet ────────────────────────────
            safe_title = f"APU-{item_counter:02d}"
            ws = wb.create_sheet(title=safe_title)

            # Column widths
            # We use a wider Col A for descriptions as requested
            for col_ltr, w in [('A',45),('B',12),('C',10),('D',10),('E',10),('F',8),('G',8),('H',12),('I',12),('J',13)]:
                ws.column_dimensions[col_ltr].width = w

            # ── Block 1: Header ───────────────────────────────────
            r = 1
            start_apu_row = r
            ws.merge_cells(f"A{r}:J{r}")
            _set_cell(ws, r, 1, "ANÁLISIS DE PRECIO UNITARIO (APU)",
                      bold=True, size=13, bg=C_HEADER, align="center")
            ws.row_dimensions[r].height = 24
            r += 1

            for label, value in [
                ("Proyecto:",    project_name),
                ("NIT/ID:",      project_nit),
                ("Actividad:",   p_name),
                ("Variante:",    v_name),
                ("Categoría:",   categoria),
                ("Unidad:",      unit),
            ]:
                _set_cell(ws, r, 1, label, bold=True, bg="D5D8DC")
                ws.merge_cells(f"B{r}:J{r}")
                _set_cell(ws, r, 2, value)
                r += 1

            # Variant Description Row
            _set_cell(ws, r, 1, "Descripción:", bold=True, bg="D5D8DC")
            ws.merge_cells(f"B{r}:J{r}")
            desc_val = v_desc or "Sin descripción"
            c_desc = _set_cell(ws, r, 2, desc_val)
            c_desc.alignment = Alignment(wrap_text=True, vertical="center")
            
            # Calculate Height: Col A-J is roughly 100 characters wide total.
            # B-J is ~90. 
            lines = max(1, len(desc_val) // 90 + desc_val.count('\n') + 1)
            ws.row_dimensions[r].height = lines * 15 + 10 
            r += 2 # Extra space

            # ── Matrix items ──────────────────────────────────────
            matrix_query = """
                SELECT i.id, i.nombre, i.tipo_item,
                       COALESCE(i.presentacion,'Und'),
                       COALESCE(i.contenido_presentacion,1.0),
                       COALESCE(i.precio_venta,0.0),
                       COALESCE(i.descuento_porcentaje,0.0),
                       COALESCE(i.desperdicio_porcentaje,0.0),
                       m.rendimiento
                FROM matriz_apu m
                JOIN insumos i ON m.insumo_id = i.id
                WHERE m.variante_id = ?
                ORDER BY CASE i.tipo_item 
                    WHEN 'Equipo' THEN 1 
                    WHEN 'Material' THEN 2 
                    WHEN 'Transporte' THEN 3 
                    WHEN 'Mano de Obra' THEN 4 
                    ELSE 5 END
            """
            matrix = engine_db.execute_query(matrix_query, (v_id,))

            total_directo = 0.0
            section_totals = {"Material": 0.0, "Equipo": 0.0, "Mano de Obra": 0.0, "Transporte": 0.0}
            current_section = None
            section_colors = {"Material": C_MAT, "Equipo": C_EQ, "Mano de Obra": C_MO, "Transporte": "F9E79F"}

            # Group items by category to write specific headers
            from collections import defaultdict
            grouped_items = defaultdict(list)
            for m_item in (matrix or []):
                grouped_items[m_item[2]].append(m_item)

            # Execution order as per user image or typical APU
            ordered_cats = ["Equipo", "Material", "Transporte", "Mano de Obra"]
            cat_counter = 1

            for cat in ordered_cats:
                items = grouped_items.get(cat, [])
                if not items: continue

                # Section Title
                ws.merge_cells(f"A{r}:J{r}")
                _set_cell(ws, r, 1, f"{cat_counter}. {cat.upper()}", bold=True, size=11)
                r += 1

                # Specific Headers and Fill Logic
                if cat == "Equipo":
                    # Headers: Desc(A-B), Marca(C-E), Tipo(F-G), Tarifa(H), Rend(I), V/Unit(J)
                    h_labels = ["Descripción", "Marca", "Tipo", "Tarifa/hora", "Rendimiento", "V/Unitario"]
                    h_cols   = [1, 3, 6, 8, 9, 10]
                    ws.merge_cells(f"A{r}:B{r}")
                    ws.merge_cells(f"C{r}:E{r}")
                    ws.merge_cells(f"F{r}:G{r}")
                    for lbl, col in zip(h_labels, h_cols):
                        _set_cell(ws, r, col, lbl, bold=True, bg=C_HEADER, align="center", border=_thin_border())
                    r += 1

                    for ins_id, nombre, tipo, pres, cont, p_v, dto, desp, rend_m2 in items:
                        rend_m2 = rend_m2 or 1.0
                        p_real  = p_v * (1 - dto / 100.0)
                        costo_m2 = p_v / max(rend_m2, 0.001)
                        total_directo += costo_m2
                        section_totals[cat] += costo_m2

                        ws.merge_cells(f"A{r}:B{r}")
                        ws.merge_cells(f"C{r}:E{r}")
                        ws.merge_cells(f"F{r}:G{r}")
                        _set_cell(ws, r, 1, nombre, border=_thin_border())
                        _set_cell(ws, r, 3, "", border=_thin_border()) # Marca
                        _set_cell(ws, r, 6, "", border=_thin_border()) # Tipo
                        _set_cell(ws, r, 8, p_v, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 9, rend_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 10, costo_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        r += 1

                elif cat == "Material":
                    # Headers: Descripción (A-G), Unidad (H), Cantidad (I), V/Unitario (J)
                    # User: "DESPUES DE UNIDAD EN MATERIAL HAY MUCHAS CELDAS VACIAS... MEJOR HACER MERGE"
                    # We'll merge A-F for Desc, G for Unidad, H for Price, I for Cant, J for V/Unit
                    h_labels = ["Descripción", "Unidad", "Precio unit.", "Cantidad", "V/Unitario"]
                    h_cols   = [1, 7, 8, 9, 10]
                    ws.merge_cells(f"A{r}:F{r}")
                    for lbl, col in zip(h_labels, h_cols):
                        _set_cell(ws, r, col, lbl, bold=True, bg=C_HEADER, align="center", border=_thin_border())
                    r += 1

                    for ins_id, nombre, tipo, pres, cont, p_v, dto, desp, rend_m2 in items:
                        rend_m2 = rend_m2 or 1.0
                        cont = cont or 1.0
                        p_real = p_v * (1 - dto / 100.0)
                        cant_m2 = ((1.0 / max(rend_m2, 0.001)) * (1 + desp / 100.0)) / max(cont, 0.001)
                        costo_m2 = p_v * cant_m2
                        total_directo += costo_m2
                        section_totals[cat] += costo_m2

                        ws.merge_cells(f"A{r}:F{r}")
                        _set_cell(ws, r, 1, nombre, border=_thin_border())
                        _set_cell(ws, r, 7, pres, align="center", border=_thin_border())
                        _set_cell(ws, r, 8, p_v, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 9, cant_m2, number_fmt='0.00000', align="center", border=_thin_border())
                        _set_cell(ws, r, 10, costo_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        r += 1

                elif cat == "Transporte":
                    # Headers: Material(A-C), Volumen(D), Distancia(E), M3-Ton/Km(F-G), Tarifa(H), V/Unit(J)
                    h_labels = ["Material", "Volumen/peso", "Distancia", "M3-Ton / Km", "Tarifa", "V/Unitario"]
                    h_cols   = [1, 4, 5, 6, 8, 10]
                    ws.merge_cells(f"A{r}:C{r}")
                    ws.merge_cells(f"F{r}:G{r}")
                    for lbl, col in zip(h_labels, h_cols):
                        _set_cell(ws, r, col, lbl, bold=True, bg=C_HEADER, align="center", border=_thin_border())
                    r += 1

                    for ins_id, nombre, tipo, pres, cont, p_v, dto, desp, rend_m2 in items:
                        rend_m2 = rend_m2 or 1.0
                        p_real  = p_v * (1 - dto / 100.0)
                        costo_m2 = p_real / max(rend_m2, 0.001)
                        total_directo += costo_m2
                        section_totals[cat] += costo_m2

                        ws.merge_cells(f"A{r}:C{r}")
                        _set_cell(ws, r, 1, nombre, border=_thin_border())
                        _set_cell(ws, r, 4, "según pedido", align="center", border=_thin_border()) 
                        _set_cell(ws, r, 5, "variable", align="center", border=_thin_border())
                        ws.merge_cells(f"F{r}:G{r}")
                        _set_cell(ws, r, 6, "", border=_thin_border())
                        _set_cell(ws, r, 8, p_v, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 10, costo_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        r += 1

                elif cat == "Mano de Obra":
                    # Headers: Personal(A-C), Jornal(D), Prestaciones(E), Jornal/Total(F-H), Rendimiento(I), V/Unit(J)
                    h_labels = ["Personal", "Jornal", "Prestaciones", "Jornal/total", "Rendimiento", "V/Unitario"]
                    h_cols   = [1, 4, 5, 6, 9, 10]
                    ws.merge_cells(f"A{r}:C{r}")
                    ws.merge_cells(f"F{r}:H{r}")
                    for lbl, col in zip(h_labels, h_cols):
                        _set_cell(ws, r, col, lbl, bold=True, bg=C_HEADER, align="center", border=_thin_border())
                    r += 1

                    for ins_id, nombre, tipo, pres, cont, p_v, dto, desp, rend_m2 in items:
                        rend_m2 = rend_m2 or 1.0
                        p_venta = p_v 
                        total_jornal = p_venta * 1.4
                        costo_m2 = p_venta * 1.4 / max(rend_m2, 0.001)
                        total_directo += costo_m2
                        section_totals[cat] += costo_m2

                        ws.merge_cells(f"A{r}:C{r}")
                        _set_cell(ws, r, 1, nombre, border=_thin_border())
                        _set_cell(ws, r, 4, p_venta, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 5, 0.40, number_fmt='0.00', align="center", border=_thin_border())
                        ws.merge_cells(f"F{r}:H{r}")
                        _set_cell(ws, r, 6, total_jornal, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 9, rend_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        _set_cell(ws, r, 10, costo_m2, number_fmt='#,##0.00', align="center", border=_thin_border())
                        r += 1

                # Subtotal for section
                ws.merge_cells(f"A{r}:I{r}")
                _set_cell(ws, r, 1, f"Sub-total {cat}", bold=True, align="right")
                _set_cell(ws, r, 10, section_totals.get(cat, 0.0), bold=True, number_fmt='#,##0.00', border=_thin_border())
                r += 2 
                cat_counter += 1

            # ── Final Summary ──────────────────
            r += 1
            ws.merge_cells(f"A{r}:I{r}")
            _set_cell(ws, r, 1, "COSTO DIRECTO TOTAL", bold=True, size=11, bg=C_TOTAL, align="right")
            _set_cell(ws, r, 10, total_directo, bold=True, size=11, bg=C_TOTAL, number_fmt='#,##0.00')
            ws.row_dimensions[r].height = 22
            
            # APPLY THICK BORDER TO THE WHOLE BLOCK
            s_med = Side(style="medium")
            # Left edge
            for row in range(start_apu_row, r + 1):
                c = ws.cell(row=row, column=1)
                c.border = Border(left=s_med, top=c.border.top, right=c.border.right, bottom=c.border.bottom)
            # Right edge
            for row in range(start_apu_row, r + 1):
                c = ws.cell(row=row, column=10)
                c.border = Border(right=s_med, top=c.border.top, left=c.border.left, bottom=c.border.bottom)
            # Top edge
            for col in range(1, 11):
                c = ws.cell(row=start_apu_row, column=col)
                c.border = Border(top=s_med, left=c.border.left, right=c.border.right, bottom=c.border.bottom)
            # Bottom edge
            for col in range(1, 11):
                c = ws.cell(row=r, column=col)
                c.border = Border(bottom=s_med, left=c.border.left, right=c.border.right, top=c.border.top)

            r += 2

            # ── Update index sheet ────────────────────────────────
            _set_cell(ws_index, idx_row, 1, item_counter, align="center")
            _set_cell(ws_index, idx_row, 2, p_name)
            _set_cell(ws_index, idx_row, 3, unit, align="center")
            _set_cell(ws_index, idx_row, 4, total_directo, number_fmt='#,##0.00')
            _set_cell(ws_index, idx_row, 5, total_directo, number_fmt='#,##0.00') # Fixed Technical
            grand_total += total_directo * qty 
            idx_row += 1
            item_counter += 1

        # ── Index grand total ─────────────────────────────────────
        ws_index.merge_cells(f"A{idx_row}:C{idx_row}")
        _set_cell(ws_index, idx_row, 1, "TOTAL GENERAL PRESUPUESTO", bold=True, bg=C_TOTAL, align="right")
        _set_cell(ws_index, idx_row, 5, grand_total, bold=True, bg=C_TOTAL, number_fmt='#,##0.00')

        wb.save(output_path)
        return output_path
